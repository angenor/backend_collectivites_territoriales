"""
Endpoints pour l'import de données via Excel
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Dict, Any
import openpyxl
from io import BytesIO
from uuid import UUID
from datetime import datetime

from app.database import get_db
from app.api.deps import get_current_active_user
from app.models.revenus import Revenu as RevenuModel, Rubrique as RubriqueModel, Periode as PeriodeModel
from app.models.geographie import Commune as CommuneModel
from app.models.utilisateurs import Utilisateur

router = APIRouter()


def detect_header_row(worksheet, max_rows=10) -> int:
    """
    Détecte automatiquement la ligne d'en-tête dans une feuille Excel.
    Cherche la première ligne contenant des valeurs texte significatives.
    """
    for row_idx in range(1, min(max_rows + 1, worksheet.max_row + 1)):
        row_values = [cell.value for cell in worksheet[row_idx]]
        # Vérifier si la ligne contient au moins 3 valeurs non-vides
        non_empty = [v for v in row_values if v is not None and str(v).strip()]
        if len(non_empty) >= 3:
            # Vérifier si ce sont principalement des strings (en-têtes)
            text_count = sum(1 for v in non_empty if isinstance(v, str))
            if text_count / len(non_empty) >= 0.7:  # Au moins 70% de texte
                return row_idx
    return 1  # Par défaut, première ligne


def parse_excel_file(file_content: bytes) -> Dict[str, Any]:
    """
    Parse un fichier Excel et retourne les informations de structure.
    """
    workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

    sheets_info = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        # Détecter la ligne d'en-tête
        header_row = detect_header_row(worksheet)

        # Extraire les en-têtes
        headers = []
        for cell in worksheet[header_row]:
            value = cell.value
            if value is not None:
                headers.append(str(value).strip())
            else:
                headers.append(f"Colonne_{len(headers) + 1}")

        # Compter les lignes de données (après l'en-tête)
        data_rows = 0
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            row_values = [cell.value for cell in worksheet[row_idx]]
            # Considérer la ligne comme données si au moins une cellule non-vide
            if any(v is not None and str(v).strip() for v in row_values):
                data_rows += 1

        sheets_info.append({
            "name": sheet_name,
            "header_row": header_row,
            "headers": headers,
            "data_rows": data_rows,
            "total_rows": worksheet.max_row,
            "total_columns": len(headers)
        })

    return {
        "sheets": sheets_info,
        "total_sheets": len(sheets_info)
    }


def extract_data_from_excel(file_content: bytes, sheet_name: str, header_row: int, column_mapping: Dict[str, str]) -> List[Dict[str, Any]]:
    """
    Extrait les données d'une feuille Excel selon un mapping de colonnes.

    column_mapping format: {
        "excel_column_name": "database_field_name",
        ...
    }
    """
    workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Feuille '{sheet_name}' introuvable")

    worksheet = workbook[sheet_name]

    # Obtenir les en-têtes
    headers = []
    for cell in worksheet[header_row]:
        value = cell.value
        if value is not None:
            headers.append(str(value).strip())
        else:
            headers.append(f"Colonne_{len(headers) + 1}")

    # Créer un mapping index -> nom de champ
    column_indexes = {}
    for excel_col, db_field in column_mapping.items():
        try:
            col_index = headers.index(excel_col)
            column_indexes[col_index] = db_field
        except ValueError:
            # Colonne non trouvée dans Excel
            continue

    # Extraire les données
    data = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        row = worksheet[row_idx]
        row_data = {}

        for col_idx, db_field in column_indexes.items():
            if col_idx < len(row):
                cell_value = row[col_idx].value
                # Nettoyer la valeur
                if cell_value is not None:
                    if isinstance(cell_value, str):
                        cell_value = cell_value.strip()
                    row_data[db_field] = cell_value

        # Ne garder que les lignes avec au moins un champ rempli
        if row_data:
            data.append(row_data)

    return data


@router.post("/upload/analyze", summary="Analyser un fichier Excel")
async def analyze_excel(
    file: UploadFile = File(...),
    current_user: Utilisateur = Depends(get_current_active_user)
):
    """
    Analyse la structure d'un fichier Excel sans importer les données.
    Retourne les feuilles, en-têtes et un aperçu de la structure.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Le fichier doit être au format Excel (.xlsx ou .xls)"
        )

    try:
        content = await file.read()
        structure = parse_excel_file(content)

        return {
            "filename": file.filename,
            "structure": structure,
            "message": "Fichier analysé avec succès"
        }
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Erreur lors de l'analyse du fichier: {str(e)}"
        )


@router.post("/upload/preview", summary="Prévisualiser l'import")
async def preview_import(
    file: UploadFile = File(...),
    sheet_name: str = "Sheet1",
    header_row: int = 1,
    current_user: Utilisateur = Depends(get_current_active_user)
):
    """
    Prévisualise les données avant import.
    Lit les premières lignes du fichier Excel.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Le fichier doit être au format Excel (.xlsx ou .xls)"
        )

    try:
        content = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)

        if sheet_name not in workbook.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"Feuille '{sheet_name}' introuvable. Feuilles disponibles: {', '.join(workbook.sheetnames)}"
            )

        worksheet = workbook[sheet_name]

        # Obtenir les en-têtes
        headers = []
        for cell in worksheet[header_row]:
            value = cell.value
            headers.append(str(value).strip() if value is not None else "")

        # Lire les 10 premières lignes de données
        preview_data = []
        for row_idx in range(header_row + 1, min(header_row + 11, worksheet.max_row + 1)):
            row = worksheet[row_idx]
            row_data = {}
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = cell.value
            if any(row_data.values()):  # Ne garder que les lignes non-vides
                preview_data.append(row_data)

        return {
            "filename": file.filename,
            "sheet_name": sheet_name,
            "header_row": header_row,
            "headers": headers,
            "preview_rows": preview_data,
            "total_rows": worksheet.max_row - header_row
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Erreur lors de la prévisualisation: {str(e)}"
        )


@router.post("/upload/import-revenus", summary="Importer les revenus depuis Excel")
async def import_revenus(
    file: UploadFile = File(...),
    sheet_name: str = "Sheet1",
    header_row: int = 1,
    commune_id: UUID = None,
    periode_id: UUID = None,
    db: Session = Depends(get_db),
    current_user: Utilisateur = Depends(get_current_active_user)
):
    """
    Importe les données de revenus depuis un fichier Excel.

    Le fichier Excel doit contenir au minimum les colonnes:
    - Code rubrique (pour identifier la rubrique)
    - Montant ou Montant prévu

    Optionnel:
    - Commune (si non fournie en paramètre)
    - Période (si non fournie en paramètre)
    - Observations
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Le fichier doit être au format Excel (.xlsx ou .xls)"
        )

    try:
        content = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)

        if sheet_name not in workbook.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"Feuille '{sheet_name}' introuvable"
            )

        worksheet = workbook[sheet_name]

        # Obtenir les en-têtes et détecter les colonnes
        headers = {}
        for idx, cell in enumerate(worksheet[header_row]):
            if cell.value:
                col_name = str(cell.value).strip().lower()
                headers[col_name] = idx

        # Mapping des colonnes (flexible)
        def find_column(possible_names):
            for name in possible_names:
                if name.lower() in headers:
                    return headers[name.lower()]
            return None

        code_rubrique_idx = find_column(['code', 'code rubrique', 'rubrique', 'code_rubrique'])
        montant_idx = find_column(['montant', 'montant reel', 'montant réel', 'montant_reel'])
        montant_prevu_idx = find_column(['montant prevu', 'montant prévu', 'budget', 'montant_prevu'])
        commune_idx = find_column(['commune', 'nom commune', 'commune_id'])
        periode_idx = find_column(['periode', 'période', 'periode_id'])
        observations_idx = find_column(['observations', 'observation', 'remarques', 'commentaires'])

        if code_rubrique_idx is None:
            raise HTTPException(
                status_code=400,
                detail="Colonne 'Code rubrique' introuvable. Colonnes disponibles: " + ", ".join(headers.keys())
            )

        if montant_idx is None and montant_prevu_idx is None:
            raise HTTPException(
                status_code=400,
                detail="Au moins une colonne 'Montant' ou 'Montant prévu' est requise"
            )

        # Importer les données
        imported_count = 0
        errors = []

        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            row = worksheet[row_idx]

            try:
                # Extraire le code rubrique
                code_rubrique = row[code_rubrique_idx].value
                if not code_rubrique:
                    continue  # Ignorer les lignes vides

                code_rubrique = str(code_rubrique).strip()

                # Chercher la rubrique
                rubrique = db.query(RubriqueModel).filter(RubriqueModel.code == code_rubrique).first()
                if not rubrique:
                    errors.append({
                        "row": row_idx,
                        "error": f"Rubrique avec code '{code_rubrique}' introuvable"
                    })
                    continue

                # Déterminer commune et période
                current_commune_id = commune_id
                current_periode_id = periode_id

                if commune_idx is not None and not current_commune_id:
                    commune_name = row[commune_idx].value
                    if commune_name:
                        commune = db.query(CommuneModel).filter(CommuneModel.nom.ilike(f"%{commune_name}%")).first()
                        if commune:
                            current_commune_id = commune.id

                if periode_idx is not None and not current_periode_id:
                    periode_code = row[periode_idx].value
                    if periode_code:
                        periode = db.query(PeriodeModel).filter(PeriodeModel.code == str(periode_code)).first()
                        if periode:
                            current_periode_id = periode.id

                if not current_commune_id or not current_periode_id:
                    errors.append({
                        "row": row_idx,
                        "error": "Commune ou période manquante"
                    })
                    continue

                # Extraire les montants
                montant = 0
                if montant_idx is not None and row[montant_idx].value is not None:
                    try:
                        montant = float(row[montant_idx].value)
                    except (ValueError, TypeError):
                        montant = 0

                montant_prevu = None
                if montant_prevu_idx is not None and row[montant_prevu_idx].value is not None:
                    try:
                        montant_prevu = float(row[montant_prevu_idx].value)
                    except (ValueError, TypeError):
                        pass

                # Observations
                observations = None
                if observations_idx is not None and row[observations_idx].value is not None:
                    observations = str(row[observations_idx].value)

                # Vérifier si l'entrée existe déjà
                existing = db.query(RevenuModel).filter(
                    RevenuModel.commune_id == current_commune_id,
                    RevenuModel.rubrique_id == rubrique.id,
                    RevenuModel.periode_id == current_periode_id
                ).first()

                if existing:
                    # Mettre à jour
                    existing.montant = montant
                    if montant_prevu is not None:
                        existing.montant_prevu = montant_prevu
                    if observations:
                        existing.observations = observations
                    existing.updated_by = current_user.id
                else:
                    # Créer nouveau
                    revenu = RevenuModel(
                        commune_id=current_commune_id,
                        rubrique_id=rubrique.id,
                        periode_id=current_periode_id,
                        montant=montant,
                        montant_prevu=montant_prevu,
                        observations=observations,
                        created_by=current_user.id,
                        updated_by=current_user.id
                    )
                    db.add(revenu)

                imported_count += 1

            except Exception as e:
                errors.append({
                    "row": row_idx,
                    "error": str(e)
                })

        # Commit tous les changements
        db.commit()

        return {
            "filename": file.filename,
            "imported_count": imported_count,
            "error_count": len(errors),
            "errors": errors[:50],  # Limiter à 50 erreurs
            "message": f"{imported_count} lignes importées avec succès"
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Erreur lors de l'import: {str(e)}"
        )
