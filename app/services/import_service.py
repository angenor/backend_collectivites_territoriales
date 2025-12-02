"""
Import service for parsing and importing Excel data.
Handles administrative account data import from Excel files.
"""

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.comptabilite import (
    DonneesDepenses,
    DonneesRecettes,
    Exercice,
    PlanComptable,
)
from app.models.geographie import Commune
from app.models.enums import TypeMouvement


@dataclass
class ImportError:
    """Represents an import error."""
    row: int
    column: str
    message: str
    value: Any = None


@dataclass
class ImportResult:
    """Result of an import operation."""
    success: bool
    recettes_imported: int = 0
    depenses_imported: int = 0
    recettes_updated: int = 0
    depenses_updated: int = 0
    errors: list[ImportError] = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


def parse_decimal(value: Any, default: Decimal = Decimal("0.00")) -> Decimal:
    """Parse a value to Decimal, returning default if invalid."""
    if value is None or value == "" or value == "-":
        return default
    try:
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        if isinstance(value, str):
            # Remove spaces and replace comma with dot
            cleaned = value.replace(" ", "").replace(",", ".")
            return Decimal(cleaned)
        return default
    except (InvalidOperation, ValueError):
        return default


class ExcelImportService:
    """Service for importing data from Excel files."""

    # Expected column mappings for recettes sheet
    RECETTES_COLUMNS = {
        "code": 0,  # A
        "intitule": 1,  # B
        "budget_primitif": 2,  # C
        "budget_additionnel": 3,  # D
        "modifications": 4,  # E
        "previsions_definitives": 5,  # F
        "or_admis": 6,  # G
        "recouvrement": 7,  # H
        "reste_a_recouvrer": 8,  # I
    }

    # Expected column mappings for depenses sheet
    DEPENSES_COLUMNS = {
        "code": 0,  # A
        "intitule": 1,  # B
        "budget_primitif": 2,  # C
        "budget_additionnel": 3,  # D
        "modifications": 4,  # E
        "previsions_definitives": 5,  # F
        "engagement": 6,  # G
        "mandat_admis": 7,  # H
        "paiement": 8,  # I
        "reste_a_payer": 9,  # J
    }

    def __init__(self, db: Session):
        self.db = db
        self._plan_comptable_cache: dict[str, PlanComptable] = {}

    def _get_plan_comptable(self, code: str) -> Optional[PlanComptable]:
        """Get PlanComptable by code, using cache."""
        if code not in self._plan_comptable_cache:
            compte = (
                self.db.query(PlanComptable)
                .filter(PlanComptable.code == code)
                .first()
            )
            self._plan_comptable_cache[code] = compte
        return self._plan_comptable_cache.get(code)

    def validate_file(
        self,
        file_content: bytes,
        commune_id: int,
        exercice_id: int,
    ) -> ImportResult:
        """
        Validate an Excel file without importing.

        Returns validation result with any errors found.
        """
        result = ImportResult(success=True)

        try:
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            result.success = False
            result.errors.append(
                ImportError(row=0, column="", message=f"Fichier Excel invalide: {str(e)}")
            )
            return result

        # Check required sheets
        sheet_names = [s.lower() for s in wb.sheetnames]
        has_recettes = any("recette" in name for name in sheet_names)
        has_depenses = any("dépense" in name or "depense" in name for name in sheet_names)

        if not has_recettes and not has_depenses:
            result.success = False
            result.errors.append(
                ImportError(
                    row=0,
                    column="",
                    message="Le fichier doit contenir au moins une feuille 'Recettes' ou 'Dépenses'",
                )
            )
            return result

        # Validate commune
        commune = self.db.query(Commune).filter(Commune.id == commune_id).first()
        if not commune:
            result.success = False
            result.errors.append(
                ImportError(row=0, column="", message="Commune non trouvée")
            )
            return result

        # Validate exercice
        exercice = self.db.query(Exercice).filter(Exercice.id == exercice_id).first()
        if not exercice:
            result.success = False
            result.errors.append(
                ImportError(row=0, column="", message="Exercice non trouvé")
            )
            return result

        if exercice.cloture:
            result.success = False
            result.errors.append(
                ImportError(
                    row=0, column="", message="Impossible d'importer dans un exercice clôturé"
                )
            )
            return result

        # Validate recettes sheet
        for sheet in wb.sheetnames:
            if "recette" in sheet.lower():
                ws = wb[sheet]
                errors = self._validate_recettes_sheet(ws)
                result.errors.extend(errors)

        # Validate depenses sheet
        for sheet in wb.sheetnames:
            if "dépense" in sheet.lower() or "depense" in sheet.lower():
                ws = wb[sheet]
                errors = self._validate_depenses_sheet(ws)
                result.errors.extend(errors)

        if result.errors:
            result.success = False

        wb.close()
        return result

    def _validate_recettes_sheet(self, ws: Worksheet) -> list[ImportError]:
        """Validate recettes sheet structure and data."""
        errors = []

        # Skip header rows (typically first 3-5 rows)
        start_row = self._find_data_start_row(ws)
        if start_row == 0:
            errors.append(
                ImportError(
                    row=0,
                    column="",
                    message="Impossible de trouver le début des données dans la feuille Recettes",
                )
            )
            return errors

        # Validate data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            code = row[self.RECETTES_COLUMNS["code"]].value
            if code is None or str(code).strip() == "":
                continue  # Skip empty rows

            code = str(code).strip()

            # Check if code exists in plan comptable
            compte = self._get_plan_comptable(code)
            if not compte:
                errors.append(
                    ImportError(
                        row=row_idx,
                        column="A",
                        message=f"Code comptable inconnu: {code}",
                        value=code,
                    )
                )
                continue

            # Validate it's a receipt code
            if compte.type_mouvement != TypeMouvement.RECETTE:
                errors.append(
                    ImportError(
                        row=row_idx,
                        column="A",
                        message=f"Le code {code} n'est pas un code de recette",
                        value=code,
                    )
                )

        return errors

    def _validate_depenses_sheet(self, ws: Worksheet) -> list[ImportError]:
        """Validate depenses sheet structure and data."""
        errors = []

        start_row = self._find_data_start_row(ws)
        if start_row == 0:
            errors.append(
                ImportError(
                    row=0,
                    column="",
                    message="Impossible de trouver le début des données dans la feuille Dépenses",
                )
            )
            return errors

        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            code = row[self.DEPENSES_COLUMNS["code"]].value
            if code is None or str(code).strip() == "":
                continue

            code = str(code).strip()

            compte = self._get_plan_comptable(code)
            if not compte:
                errors.append(
                    ImportError(
                        row=row_idx,
                        column="A",
                        message=f"Code comptable inconnu: {code}",
                        value=code,
                    )
                )
                continue

            if compte.type_mouvement != TypeMouvement.DEPENSE:
                errors.append(
                    ImportError(
                        row=row_idx,
                        column="A",
                        message=f"Le code {code} n'est pas un code de dépense",
                        value=code,
                    )
                )

        return errors

    def _find_data_start_row(self, ws: Worksheet) -> int:
        """Find the row where actual data starts."""
        # Look for a row with a valid account code in column A
        for row_idx, row in enumerate(ws.iter_rows(max_row=20), start=1):
            code = row[0].value
            if code and str(code).strip():
                code_str = str(code).strip()
                # Check if it looks like an account code (numeric or alphanumeric)
                if code_str[0].isdigit() or self._get_plan_comptable(code_str):
                    return row_idx
        return 0

    def import_file(
        self,
        file_content: bytes,
        commune_id: int,
        exercice_id: int,
        update_existing: bool = True,
    ) -> ImportResult:
        """
        Import data from Excel file.

        Args:
            file_content: Excel file bytes
            commune_id: Target commune ID
            exercice_id: Target exercice ID
            update_existing: If True, update existing entries; if False, skip them

        Returns:
            ImportResult with counts and any errors
        """
        # First validate
        result = self.validate_file(file_content, commune_id, exercice_id)
        if not result.success:
            return result

        try:
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            result.success = False
            result.errors.append(
                ImportError(row=0, column="", message=f"Erreur lecture fichier: {str(e)}")
            )
            return result

        # Import recettes
        for sheet in wb.sheetnames:
            if "recette" in sheet.lower():
                ws = wb[sheet]
                counts = self._import_recettes_sheet(
                    ws, commune_id, exercice_id, update_existing
                )
                result.recettes_imported += counts[0]
                result.recettes_updated += counts[1]

        # Import depenses
        for sheet in wb.sheetnames:
            if "dépense" in sheet.lower() or "depense" in sheet.lower():
                ws = wb[sheet]
                counts = self._import_depenses_sheet(
                    ws, commune_id, exercice_id, update_existing
                )
                result.depenses_imported += counts[0]
                result.depenses_updated += counts[1]

        wb.close()

        # Commit changes
        self.db.commit()
        result.success = True

        return result

    def _import_recettes_sheet(
        self,
        ws: Worksheet,
        commune_id: int,
        exercice_id: int,
        update_existing: bool,
    ) -> tuple[int, int]:
        """Import recettes from worksheet. Returns (imported, updated) counts."""
        imported = 0
        updated = 0

        start_row = self._find_data_start_row(ws)
        if start_row == 0:
            return (0, 0)

        for row in ws.iter_rows(min_row=start_row):
            code = row[self.RECETTES_COLUMNS["code"]].value
            if code is None or str(code).strip() == "":
                continue

            code = str(code).strip()
            compte = self._get_plan_comptable(code)
            if not compte or compte.type_mouvement != TypeMouvement.RECETTE:
                continue

            # Extract values
            values = {
                "budget_primitif": parse_decimal(
                    row[self.RECETTES_COLUMNS["budget_primitif"]].value
                ),
                "budget_additionnel": parse_decimal(
                    row[self.RECETTES_COLUMNS["budget_additionnel"]].value
                ),
                "modifications": parse_decimal(
                    row[self.RECETTES_COLUMNS["modifications"]].value
                ),
                "previsions_definitives": parse_decimal(
                    row[self.RECETTES_COLUMNS["previsions_definitives"]].value
                ),
                "or_admis": parse_decimal(
                    row[self.RECETTES_COLUMNS["or_admis"]].value
                ),
                "recouvrement": parse_decimal(
                    row[self.RECETTES_COLUMNS["recouvrement"]].value
                ),
                "reste_a_recouvrer": parse_decimal(
                    row[self.RECETTES_COLUMNS["reste_a_recouvrer"]].value
                ),
            }

            # Check if entry exists
            existing = (
                self.db.query(DonneesRecettes)
                .filter(
                    DonneesRecettes.commune_id == commune_id,
                    DonneesRecettes.exercice_id == exercice_id,
                    DonneesRecettes.compte_code == code,
                )
                .first()
            )

            if existing:
                if update_existing:
                    for field, value in values.items():
                        setattr(existing, field, value)
                    updated += 1
            else:
                recette = DonneesRecettes(
                    commune_id=commune_id,
                    exercice_id=exercice_id,
                    compte_code=code,
                    **values,
                )
                self.db.add(recette)
                imported += 1

        return (imported, updated)

    def _import_depenses_sheet(
        self,
        ws: Worksheet,
        commune_id: int,
        exercice_id: int,
        update_existing: bool,
    ) -> tuple[int, int]:
        """Import depenses from worksheet. Returns (imported, updated) counts."""
        imported = 0
        updated = 0

        start_row = self._find_data_start_row(ws)
        if start_row == 0:
            return (0, 0)

        for row in ws.iter_rows(min_row=start_row):
            code = row[self.DEPENSES_COLUMNS["code"]].value
            if code is None or str(code).strip() == "":
                continue

            code = str(code).strip()
            compte = self._get_plan_comptable(code)
            if not compte or compte.type_mouvement != TypeMouvement.DEPENSE:
                continue

            values = {
                "budget_primitif": parse_decimal(
                    row[self.DEPENSES_COLUMNS["budget_primitif"]].value
                ),
                "budget_additionnel": parse_decimal(
                    row[self.DEPENSES_COLUMNS["budget_additionnel"]].value
                ),
                "modifications": parse_decimal(
                    row[self.DEPENSES_COLUMNS["modifications"]].value
                ),
                "previsions_definitives": parse_decimal(
                    row[self.DEPENSES_COLUMNS["previsions_definitives"]].value
                ),
                "engagement": parse_decimal(
                    row[self.DEPENSES_COLUMNS["engagement"]].value
                ),
                "mandat_admis": parse_decimal(
                    row[self.DEPENSES_COLUMNS["mandat_admis"]].value
                ),
                "paiement": parse_decimal(
                    row[self.DEPENSES_COLUMNS["paiement"]].value
                ),
                "reste_a_payer": parse_decimal(
                    row[self.DEPENSES_COLUMNS["reste_a_payer"]].value
                ),
            }

            existing = (
                self.db.query(DonneesDepenses)
                .filter(
                    DonneesDepenses.commune_id == commune_id,
                    DonneesDepenses.exercice_id == exercice_id,
                    DonneesDepenses.compte_code == code,
                )
                .first()
            )

            if existing:
                if update_existing:
                    for field, value in values.items():
                        setattr(existing, field, value)
                    updated += 1
            else:
                depense = DonneesDepenses(
                    commune_id=commune_id,
                    exercice_id=exercice_id,
                    compte_code=code,
                    **values,
                )
                self.db.add(depense)
                imported += 1

        return (imported, updated)
