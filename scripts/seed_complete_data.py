"""
Script de seed complet pour peupler la base de données
avec toutes les rubriques et données de test
"""

import os
import sys
from pathlib import Path

# Ajouter le répertoire parent au PYTHONPATH
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy.orm import Session
from decimal import Decimal
from datetime import date, timedelta
import random

from app.database import engine, SessionLocal
from app.models.revenus import Rubrique, Exercice, Periode, Revenu
from app.models.geographie import Commune, Departement, Region
from app.models.projets_miniers import ProjetMinier, SocieteMiniere, TypeMinerai


def parse_excel_rubriques(excel_file: str):
    """Parse le fichier Excel pour extraire les rubriques"""
    wb = openpyxl.load_workbook(excel_file, data_only=True)

    rubriques = []

    # Parser RECETTES
    print("📊 Parsing RECETTES...")
    sheet = wb['RECETTE']
    for row in sheet.iter_rows(min_row=8, values_only=True):
        code = None
        niveau = 0
        intitule = row[4] if len(row) > 4 else None

        # Déterminer le code et le niveau
        try:
            if row[1] and isinstance(row[1], (int, float)):  # Niveau 1 (colonne B)
                code = str(int(row[1]))
                niveau = 1
            elif row[2] and isinstance(row[2], (int, float)):  # Niveau 2 (colonne C)
                code = str(int(row[2]))
                niveau = 2
            elif row[3] and isinstance(row[3], (int, float)):  # Niveau 3 (colonne D)
                code = str(int(row[3]))
                niveau = 3
        except (ValueError, TypeError):
            continue

        if code and intitule and isinstance(intitule, str):
            rubriques.append({
                'code': f'R{code}',
                'nom': intitule.strip(),
                'niveau': niveau,
                'type': 'recette',
                'est_calculee': niveau <= 2,  # Niveaux 1 et 2 sont calculés
                'afficher_total': True
            })

    print(f"   ✓ {len(rubriques)} recettes trouvées")

    # Parser DEPENSES
    print("📊 Parsing DEPENSES...")
    sheet = wb['DEPENSES']
    count_before = len(rubriques)
    for row in sheet.iter_rows(min_row=10, values_only=True):
        code = None
        niveau = 0
        intitule = row[4] if len(row) > 4 else None

        # Déterminer le code et le niveau
        try:
            if row[1] and isinstance(row[1], (int, float)):  # Niveau 1 (colonne B)
                code = str(int(row[1]))
                niveau = 1
            elif row[2] and isinstance(row[2], (int, float)):  # Niveau 2 (colonne C)
                code = str(int(row[2]))
                niveau = 2
            elif row[3] and isinstance(row[3], (int, float)):  # Niveau 3 (colonne D)
                code = str(int(row[3]))
                niveau = 3
        except (ValueError, TypeError):
            continue

        if code and intitule and isinstance(intitule, str):
            rubriques.append({
                'code': f'D{code}',
                'nom': intitule.strip(),
                'niveau': niveau,
                'type': 'depense',
                'est_calculee': niveau <= 2,  # Niveaux 1 et 2 sont calculés
                'afficher_total': True
            })

    print(f"   ✓ {len(rubriques) - count_before} dépenses trouvées")

    # Ajouter quelques rubriques de solde
    rubriques.extend([
        {
            'code': 'S001',
            'nom': 'SOLDE DE FONCTIONNEMENT',
            'niveau': 1,
            'type': 'solde',
            'est_calculee': True,
            'afficher_total': True
        },
        {
            'code': 'S002',
            'nom': 'SOLDE GLOBAL',
            'niveau': 1,
            'type': 'solde',
            'est_calculee': True,
            'afficher_total': True
        }
    ])

    # Éliminer les doublons (garder la première occurrence)
    rubriques_uniques = {}
    for rubrique in rubriques:
        code = rubrique['code']
        if code not in rubriques_uniques:
            rubriques_uniques[code] = rubrique

    rubriques_finales = list(rubriques_uniques.values())

    print(f"\n✅ Total: {len(rubriques_finales)} rubriques uniques à créer ({len(rubriques) - len(rubriques_finales)} doublons éliminés)")
    return rubriques_finales


def seed_rubriques(db: Session, rubriques_data: list):
    """Créer les rubriques dans la base de données"""
    print("\n🌱 Création des rubriques...")

    # Supprimer les rubriques existantes
    db.query(Rubrique).delete()
    db.commit()

    rubriques_created = []
    for data in rubriques_data:
        rubrique = Rubrique(
            code=data['code'],
            nom=data['nom'],
            niveau=data['niveau'],
            type=data['type'],
            est_calculee=data['est_calculee'],
            afficher_total=data['afficher_total'],
            ordre=len(rubriques_created) + 1,
            actif=True
        )
        db.add(rubrique)
        rubriques_created.append(rubrique)

    db.commit()
    print(f"   ✓ {len(rubriques_created)} rubriques créées")
    return rubriques_created


def seed_exercices_periodes(db: Session):
    """Créer des exercices et périodes"""
    print("\n🌱 Création des exercices et périodes...")

    # Supprimer les exercices existants (cascade supprimera aussi les périodes)
    db.query(Exercice).delete()
    db.commit()

    exercices = []
    periodes = []

    for annee in [2023, 2024, 2025]:
        # Créer l'exercice
        exercice = Exercice(
            annee=annee,
            date_debut=date(annee, 1, 1),
            date_fin=date(annee, 12, 31),
            statut='ouvert' if annee == 2025 else 'cloturé',
            actif=True
        )
        db.add(exercice)
        db.flush()
        exercices.append(exercice)

        # Créer les périodes
        periodes_data = [
            ('T1', f'Trimestre 1 {annee}', date(annee, 1, 1), date(annee, 3, 31), 'trimestriel', 1),
            ('T2', f'Trimestre 2 {annee}', date(annee, 4, 1), date(annee, 6, 30), 'trimestriel', 2),
            ('T3', f'Trimestre 3 {annee}', date(annee, 7, 1), date(annee, 9, 30), 'trimestriel', 3),
            ('T4', f'Trimestre 4 {annee}', date(annee, 10, 1), date(annee, 12, 31), 'trimestriel', 4),
            ('S1', f'Semestre 1 {annee}', date(annee, 1, 1), date(annee, 6, 30), 'semestriel', 5),
            ('S2', f'Semestre 2 {annee}', date(annee, 7, 1), date(annee, 12, 31), 'semestriel', 6),
            ('ANNUEL', f'Année {annee}', date(annee, 1, 1), date(annee, 12, 31), 'annuel', 7),
        ]

        for code, nom, debut, fin, type_periode, ordre in periodes_data:
            periode = Periode(
                code=f'{code}_{annee}',
                nom=nom,
                exercice_id=exercice.id,
                date_debut=debut,
                date_fin=fin,
                type_periode=type_periode,
                ordre=ordre,
                actif=True
            )
            db.add(periode)
            periodes.append(periode)

    db.commit()
    print(f"   ✓ {len(exercices)} exercices créés")
    print(f"   ✓ {len(periodes)} périodes créées")
    return exercices, periodes


def seed_projets_miniers(db: Session, communes: list):
    """Créer des projets miniers de test"""
    print("\n🌱 Création des projets miniers...")

    # Supprimer les données existantes
    db.query(ProjetMinier).delete()
    db.query(SocieteMiniere).delete()
    db.query(TypeMinerai).delete()
    db.commit()

    # Types de minerais
    types = [
        ('NICKEL', 'Nickel', 'Minerai de nickel'),
        ('ILMENITE', 'Ilménite', 'Minerai de titane'),
        ('COBALT', 'Cobalt', 'Minerai de cobalt'),
        ('OR', 'Or', 'Minerai d\'or'),
        ('CHROME', 'Chrome', 'Minerai de chrome'),
    ]

    types_minerais = []
    for code, nom, desc in types:
        type_m = TypeMinerai(code=code, nom=nom, description=desc, actif=True)
        db.add(type_m)
        types_minerais.append(type_m)

    db.flush()

    # Sociétés minières
    societes_data = [
        ('QMM', 'QMM S.A.', 'QIT Madagascar Minerals', 'Fort Dauphin'),
        ('AMBATOVY', 'Ambatovy S.A.', 'Ambatovy Nickel', 'Toamasina'),
        ('KRAOMA', 'KRAOMA S.A.', 'Kraoma Chrome', 'Antsiranana'),
    ]

    societes = []
    for code, nom, raison, lieu in societes_data:
        societe = SocieteMiniere(
            code=code,
            nom=nom,
            raison_sociale=raison,
            adresse=lieu,
            actif=True
        )
        db.add(societe)
        societes.append(societe)

    db.flush()

    # Projets miniers
    projets_data = [
        ('QMM-ILM', 'QMM - Exploitation Ilménite', societes[0], types_minerais[1], communes[0]),
        ('AMB-NI', 'Ambatovy - Exploitation Nickel', societes[1], types_minerais[0], communes[1]),
        ('KRA-CHR', 'Kraoma - Exploitation Chrome', societes[2], types_minerais[4], communes[2]),
    ]

    projets = []
    for code, nom, societe, type_m, commune in projets_data:
        projet = ProjetMinier(
            code=code,
            nom=nom,
            societe_miniere_id=societe.id,
            type_minerai_id=type_m.id,
            commune_id=commune.id,
            date_debut=date(2020, 1, 1),
            statut='actif',
            actif=True
        )
        db.add(projet)
        projets.append(projet)

    db.commit()
    print(f"   ✓ {len(types_minerais)} types de minerais créés")
    print(f"   ✓ {len(societes)} sociétés créées")
    print(f"   ✓ {len(projets)} projets créés")
    return projets


def seed_revenus(db: Session, rubriques: list, communes: list, periodes: list, projets: list):
    """Créer des revenus de test avec toutes les colonnes du modèle comptable"""
    print("\n🌱 Création des revenus...")

    # Supprimer les revenus existants
    db.query(Revenu).delete()
    db.commit()

    revenus_created = 0

    # Pour chaque commune
    for commune in communes[:3]:  # Limiter à 3 communes
        print(f"   🏘️  Commune: {commune.nom}")

        # Pour chaque période annuelle des 2 dernières années
        periodes_annuelles = [p for p in periodes if p.type_periode == 'annuel'][-2:]

        for periode in periodes_annuelles:
            print(f"      📅 Période: {periode.nom}")

            # Pour chaque rubrique de niveau 3 (détaillées)
            rubriques_detail = [r for r in rubriques if r.niveau == 3 and not r.est_calculee]

            for rubrique in rubriques_detail[:50]:  # Limiter pour ne pas surcharger
                # Générer des montants réalistes selon le modèle comptable
                budget_primitif = Decimal(random.randint(100000, 10000000))
                budget_additionnel = Decimal(random.randint(0, int(float(budget_primitif) * 0.2)))
                modifications = Decimal(random.randint(-int(float(budget_primitif) * 0.1), int(float(budget_primitif) * 0.1)))
                previsions_definitives = budget_primitif + budget_additionnel + modifications

                # Taux de réalisation entre 70% et 120%
                taux = random.uniform(0.7, 1.2)

                if rubrique.type == 'recette':
                    # Pour les recettes
                    ordre_recette_admis = Decimal(int(float(previsions_definitives) * taux))
                    recouvrement = Decimal(int(float(ordre_recette_admis) * random.uniform(0.8, 1.0)))
                    reste_a_recouvrer = ordre_recette_admis - recouvrement
                    taux_realisation = (ordre_recette_admis / previsions_definitives * 100) if previsions_definitives > 0 else Decimal(0)

                    revenu = Revenu(
                        commune_id=commune.id,
                        rubrique_id=rubrique.id,
                        periode_id=periode.id,
                        projet_minier_id=projets[0].id if random.random() > 0.7 else None,
                        # Nouvelles colonnes de budget
                        budget_primitif=budget_primitif,
                        budget_additionnel=budget_additionnel,
                        modifications=modifications,
                        previsions_definitives=previsions_definitives,
                        # Colonnes recettes
                        ordre_recette_admis=ordre_recette_admis,
                        recouvrement=recouvrement,
                        reste_a_recouvrer=reste_a_recouvrer,
                        # Anciennes colonnes (backward compatibility)
                        montant=ordre_recette_admis,
                        montant_prevu=budget_primitif,
                        ecart=ordre_recette_admis - previsions_definitives,
                        taux_realisation=taux_realisation,
                        valide=True
                    )
                else:
                    # Pour les dépenses
                    engagement = Decimal(int(float(previsions_definitives) * random.uniform(0.8, 1.0)))
                    mandat_admis = Decimal(int(float(engagement) * random.uniform(0.9, 1.0)))
                    paiement = Decimal(int(float(mandat_admis) * random.uniform(0.7, 0.95)))
                    reste_a_payer = mandat_admis - paiement
                    taux_realisation = (mandat_admis / previsions_definitives * 100) if previsions_definitives > 0 else Decimal(0)

                    revenu = Revenu(
                        commune_id=commune.id,
                        rubrique_id=rubrique.id,
                        periode_id=periode.id,
                        projet_minier_id=projets[0].id if random.random() > 0.7 else None,
                        # Nouvelles colonnes de budget
                        budget_primitif=budget_primitif,
                        budget_additionnel=budget_additionnel,
                        modifications=modifications,
                        previsions_definitives=previsions_definitives,
                        # Colonnes dépenses
                        engagement=engagement,
                        mandat_admis=mandat_admis,
                        paiement=paiement,
                        reste_a_payer=reste_a_payer,
                        # Anciennes colonnes (backward compatibility)
                        montant=mandat_admis,
                        montant_prevu=budget_primitif,
                        ecart=mandat_admis - previsions_definitives,
                        taux_realisation=taux_realisation,
                        valide=True
                    )

                db.add(revenu)
                revenus_created += 1

    db.commit()
    print(f"\n   ✓ {revenus_created} revenus créés avec toutes les colonnes du modèle comptable")


def main():
    """Fonction principale"""
    print("=" * 70)
    print("🚀 SEED COMPLET DE LA BASE DE DONNÉES")
    print("=" * 70)

    # Chemin du fichier Excel
    excel_file = Path(__file__).parent.parent / 'bank' / 'cahier_des_charges' / 'Tableaux_de_Compte_Administratif.xlsx'

    if not excel_file.exists():
        print(f"❌ Fichier Excel introuvable: {excel_file}")
        return

    db = SessionLocal()

    try:
        # 1. Parser et créer les rubriques
        rubriques_data = parse_excel_rubriques(str(excel_file))
        rubriques = seed_rubriques(db, rubriques_data)

        # 2. Créer les exercices et périodes
        exercices, periodes = seed_exercices_periodes(db)

        # 3. Récupérer les communes existantes
        communes = db.query(Commune).limit(10).all()
        if not communes:
            print("❌ Aucune commune trouvée ! Veuillez d'abord créer des communes.")
            return

        print(f"\n📍 {len(communes)} communes trouvées")

        # 4. Créer les projets miniers
        projets = seed_projets_miniers(db, communes)

        # 5. Créer les revenus
        seed_revenus(db, rubriques, communes, periodes, projets)

        print("\n" + "=" * 70)
        print("✅ SEED TERMINÉ AVEC SUCCÈS!")
        print("=" * 70)
        print(f"\n📊 Résumé:")
        print(f"   • Rubriques: {len(rubriques)}")
        print(f"   • Exercices: {len(exercices)}")
        print(f"   • Périodes: {len(periodes)}")
        print(f"   • Projets miniers: {len(projets)}")
        print(f"   • Communes utilisées: {len(communes)}")
        print(f"\n💡 Vous pouvez maintenant exporter les données au format Excel!")

    except Exception as e:
        print(f"\n❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
    finally:
        db.close()


if __name__ == '__main__':
    main()
